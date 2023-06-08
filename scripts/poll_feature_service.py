import os
import sys
import six
import json
import time
import re
import pytz
import zipfile
import shutil
import smtplib
import openpyxl
import traceback
import requests
import inspect
import uuid
import pandas as pd
import numpy as np
from glob import glob
from datetime import datetime, timedelta
from sqlalchemy import create_engine

import download_feature_service as download
import import_track
import db_utils
import process_emails

# import warnings
# warnings.simplefilter("error")

pd.set_option('display.max_columns', None)  # useful for debugging
pd.options.mode.chained_assignment = None

SUBMISSION_TICKET_INTERVAL = 900 # seconds between submission of a particular user
LOG_CACHE_DAYS = 365  # amount of time to keep a log file
TIMESTAMP_FORMAT = '%Y-%m-%d %H:%M:%S'
TIMEZONE = pytz.timezone('US/Alaska')
DATA_PROCESSED = False  # keep track of whether any data were processed to be able to effectively log
# LANDING_FEE = 5.15 # per person fee for each passenger on scenic flight or dropped off. ### now reteived from DB
# Collect messages to warn either the user, the landings data steward, or the track editors. Also log all of them
MESSAGES = []  # internal messages to data stewards
EMAILS = []  # messages to submitters
# AGOL IDs of submissions that caused errors and data stewards have been warned about. Errors are dict of ID: date so that old
ERRORS = []
TRACK_JSON_FILES = [] # keep track of newly created track files
LANDINGS_FLIGHT_COLUMNS = {'landing_operator': 'operator_code',
                           'landing_tail_number': 'registration',
                           'landing_datetime': 'departure_datetime',
                           'landing_route': 'scenic_route',
                           'landing_aircraft_type': 'aircraft_type',
                           'globalid': 'agol_global_id',
                           'ticket': 'ticket',
                           'submission_method': 'submission_method',
                           'flight_id': 'flight_id',
                           'source_file': 'source_file',
                           'landing_flight_notes': 'notes',
                           'fee_per_passenger': 'fee_per_passenger'
                           }
LANDINGS_COLUMNS = {'flight_id': 'flight_id',
                    'landing_location': 'location',
                    'n_passengers': 'n_passengers',
                    'landing_type': 'landing_type',
                    'landing_justification': 'justification',
                    'landing_notes': 'notes',
                    'sort_order': 'sort_order'
                    }
# If there weren't any landings entered, but an Excel file of landings was submitted, the landings table will be empty
#   To populste it properly, all fields need to be present
AGOL_LANDING_COLUMNS = ['objectid', 'globalid', 'landing_type', 'landing_location', 'landing_justification', 'n_passengers_scenic', 'n_passengers_pickup', 'n_passengers_dropoff', 'landing_notes', 'parentglobalid',
       'CreationDate', 'Creator', 'EditDate', 'Editor']

LANDING_RECEIPT_COLUMNS = ['departure_datetime',
                           'registration',
                           'aircraft_type',
                           'scenic_1_passengers',
                           'scenic_route',
                           'scenic_1_location',
                           'dropoff_1_passengers',
                           'dropoff_1_location',
                           'pickup_1_passengers',
                           'pickup_1_location',
                           'dropoff_2_passengers',
                           'dropoff_2_location',
                           'pickup_2_passengers',
                           'pickup_2_location',
                           'dropoff_3_passengers',
                           'dropoff_3_location',
                           'pickup_3_passengers',
                           'pickup_3_location',
                           'notes',
                           'n_passengers',
                           'n_fee_passengers',
                           'fee'
                           ]


def read_json_params(params_json):
    '''
    Read and validate a parameters JSON file
    :param params_json: path to JSON file
    :return: dictionary of params
    '''
    required = pd.Series(['agol_credentials',
                          'db_credentials',
                          'track_data_stewards',
                          'landing_data_stewards',
                          'mail_sender',
                          'track_editor_url',
                          'delete_features_url',
                          'mail_server_credentials',
                          'log_dir',
                          'web_data_dir',
                          'ssl_cert',
                          'landing_receipt',
                          'admin_email_address'])
    with open(params_json) as j:
        params = json.load(j)
    missing = required.loc[~required.isin(params.keys())]
    if len(missing):
        if 'log_dir' in params.keys():
            msg = 'Invalid config JSON: {file}. It must contain all of "{required}" but "{missing}" are missing' \
                .format(file=params_json, required='", "'.join(required), missing='", "'.join(missing))
            write_log(params['log_dir'], 'unknown', datetime.now().strftime(TIMESTAMP_FORMAT), DATA_PROCESSED,
                      error=msg)
        raise ValueError(msg)

    if 'download_dir' not in required:
        required['download_dir'] = params['log_dir']

    return params


def read_logs(log_dir, script_name):
    log_info = []
    for json_path in glob(os.path.join(log_dir, '%s_log_*.json' % script_name)):
        # Try to read the file, but if the JSON isn't valid, just skip it
        timestamp = os.path.basename(json_path).replace('%s_log_' % script_name, '').replace('.json', '')
        try:
            if (datetime.now() - datetime.strptime(timestamp, '%Y%m%d%H%M%S')).days > 30:
                continue
        except:
            pass

        try:
            with open(json_path) as j:
                log_info.append(json.load(j))
        except:
            continue

    return pd.DataFrame(log_info)


def write_log(log_dir, download_dir, timestamp, data_was_downloaded, error='none'):
    log_file_path = os.path.join(log_dir, '{0}_log_{1}.json'.format(os.path.basename(__file__).replace('.py', ''),
                                                                    re.sub('\D', '', timestamp)))

    emails = [{k: v for k, v in e.items() if k != 'server'} for e in EMAILS]
    log_info = {'download_dir': download_dir,
                'data_downloaded': data_was_downloaded,
                'error': error,
                'timestamp': timestamp,
                'log_file_path': log_file_path,
                'submission_errors': ERRORS,
                'emails': emails}

    with open(log_file_path, 'w') as j:
        json.dump(log_info, j, indent=4)

    return log_file_path, log_info


def log_and_exit(log_info, write_log_args):
    ''' Helper function to delete old logs and write a new one before exiting'''
    try:
        # Delete old logs
        log_info.loc[(datetime.now() - log_info.timestamp).days > LOG_CACHE_DAYS, 'log_file_path'].apply(os.remove)
    except:
        pass
    sys.exit(write_log(*write_log_args))


def julian_date_to_datetime(julian_date):
    '''
    createReplica stores datetimes in an sqlite DB as a Julian date (https://simple.wikipedia.org/wiki/Julian_day) so
    convert to Unix Epoch Time
    :param timestamp:
    :return:
    '''

    # Python (really Unix) uses 1970-1-1 00:00:00 as the epoch (i.e., reference timestamp).
    JULIAN_UNIX_EPOCH = 2440587.5  # 1970-1-1 00:00:00 as a Julian date
    if pd.isna(julian_date) or julian_date < JULIAN_UNIX_EPOCH:
        return pd.NaT
    try:
        epoch_timestamp = round((julian_date - JULIAN_UNIX_EPOCH) * 60 * 60 * 24)  # number of seconds since Unix epoch
        return datetime(1970, 1, 1) + timedelta(
            seconds=epoch_timestamp)  # datetime.fromtimestamp((julian_date - JULIAN_UNIX_EPOCH) * 60 * 60 * 24)
    except:
        return pd.NaT


def get_ticket(group, connections):
    '''
    Return a dataframe with a ticket number assigned to all records in a new column
    '''

    submission_type = group.submission_type.iloc[0]  # should be all the same submission type
    submitter = group.submitter.iloc[0]
    submission_time = group.submission_time.min()
    conn = connections[submission_type]
    conn.execute(
        "INSERT INTO submissions (submitter, submission_time) VALUES ('%s', '%s');" % (submitter, submission_time))
    group['ticket'] = pd.read_sql("SELECT max(ticket) FROM submissions;", conn).squeeze()

    return group


def get_submitter_email(agol_user, agol_user_emails):
    '''
    Helper function to get an email address for the given AGOL account
    :param agol_user:
    :return:
    '''
    # All NPS accounts have an alternate email alias that is just their UPN and all NPS AGOl accounts are the UPN
    #   with "_nps" appended
    agol_user = agol_user.lower()
    if agol_user.endswith('_nps'):
        return agol_user.replace('_nps', '')
    # Otherwise, try to look up the user in the
    elif agol_user in agol_user_emails:
        return agol_user_emails[agol_user]
    else:
        return agol_user


def get_html_table(data, column_widths={}):
    thead_tr_style = 'style="font-weight: bold; height: 40px; background-color: #cccccc;'
    cell_style = 'style="padding-left: 20px; padding-right: 20px;"'
    thead_html = ('<tr {tr_style}"><th {cell_style}>' + '</th><th {cell_style}>'.join(data.columns) + '</th></tr>') \
        .format(tr_style=thead_tr_style, cell_style=cell_style)
    tr_bg_dark = ' background-color: #f2f2f2;'
    tr_content = data.apply(
        lambda x: ('<td {cell_style}>' + '</td><td {cell_style}>'.join(x.astype(str)) + '</td>').format(
            cell_style=cell_style),
        axis=1)
    tbody_html = ''.join(['<tr style="height: 40px; {bg_style}">{td_elements}'.format(
        bg_style=tr_bg_dark if i % 2 == 1 else '', td_elements=td_elements, cell_style=cell_style) for i, td_elements in
        tr_content.items()]) + '</tr>'

    for column_name, width in column_widths.items():
        thead_html.replace('<th {cell_style}>{column}</th>'.format(cell_style=cell_style, column=column_name),
                           '<th {cell_style}>{column}</th>'.format(
                               cell_style=cell_style.replace(';"', '; min-width:%s;"' % width), column=column_name))

    html = \
        """<table style="border-spacing: 0px;">
            <thead>
                {table_header}
            </thead>
            <tbody>
                {table_body}
            </tbody>
        </table>""".format(table_header=thead_html, table_body=tbody_html)

    return html


def get_email_html(data, start_datetime, end_datetime, column_widths={}):
    start_date_str = start_datetime.strftime('%B %d, %Y')
    start_time_str = start_datetime.strftime('%I:%M %p').lstrip('0')
    end_time_str = end_datetime.strftime('%I:%M %p').lstrip('0')
    # table_html = '%s %s' % ('<h4>Landings</h4><br>' if len(track_data) else '', get_html_table(landings_data))
    table_html = get_html_table(data, column_widths=column_widths)
    html = """
    <html>
        <head></head>
        <body>
            <p>
                Thank you for your flight data submission. Below is a receipt of the data you submitted on {start_date} between {start_time} and {end_time}.
                <br>
                <br>
            </p>
                {table}
            <p>
                <br>
                Note that any data you submitted after more than a 15 minute break will be included on a separate receipt.
                <br>
                <br>
                If you have any questions about this receipt or if you think there is an error in the data, you can reply to this message and a member of our team will get back to you as soon as possible. 
            </p>
        </body>
    </html>
    """.format(start_date=start_date_str, start_time=start_time_str, end_time=end_time_str, table=table_html)

    return html


def get_attachment_info(file_path, submissions, ext='.zip'):
    ''' Each zip file created by download_feature_service.download_data() has an accompanying JSON metadata file, so return the JSON as a dict'''
    json_meta_path = file_path.replace(ext, '.json')
    with open(json_meta_path) as j:
        metadata = json.load(j)
    engine = create_engine('sqlite:///' + metadata['sqlite_path'])

    with engine.connect() as conn:
        submission_data = pd.read_sql(
            "SELECT * FROM %s WHERE globalid='%s';" % (metadata['parent_table_name'], metadata['REL_GLOBALID']),
            conn).squeeze(axis=0)

    submission_data['ticket'] = submissions.loc[submissions.globalid == metadata['REL_GLOBALID'], 'ticket'].squeeze()

    return pd.concat([pd.Series(metadata), submission_data])


def track_to_json(gdf, attachment_info):
    unserializable_fields = ~gdf.dtypes.astype(str).isin(['int64', 'int32', 'float64', 'object', 'geometry'])
    gdf.loc[:, unserializable_fields] = gdf.loc[:, unserializable_fields].astype(str)
    geojson_strs = {}  # seg_id: json.loads(df.to_json()) for seg_id, df in gdf.groupby('segment_id')}
    for seg_id, df in gdf.groupby('segment_id'):  # .apply(lambda df: df.index.min())
        df['min_index'] = df.index.min()
        df['point_index'] = df.index
        geojson_strs[seg_id] = json.loads(df.to_json())

    return {'geojsons': geojson_strs, 'track_info': attachment_info.astype(str).to_dict()}


def delete_from_feature_service(agol_ids, token, service_url, ssl_cert):
    agol_id_str = ("'%s'" % "', '".join(agol_ids)).upper()
    query_params = {'f': 'json',
                    'token': token,
                    'where': "upper(globalid) IN (%s)" % agol_id_str,
                    'returnDeleteResults': True}
    query_response = requests.post('%s/0/deleteFeatures' % service_url, params=query_params, verify=ssl_cert)
    download.check_http_error('deleting %s' % agol_id_str, query_response)
    result = query_response.json()
    result_df = pd.DataFrame(result['deleteResults'])

    failed_deletes = result_df.loc[~result_df.success, 'objectId']

    return failed_deletes


def get_traceback_frame():
    ''' Traverse the traceback and get the frame of the actual script that caused the last error'''
    traceback_exc = traceback.TracebackException(*sys.exc_info())
    current_script_dir = os.path.dirname(os.path.realpath(__file__))
    for frame in traceback_exc.stack[::-1]:  # traverse in reverse order to get topmost first
        # this only works for errors caused by scripts in the same dir
        if os.path.realpath(os.path.dirname(frame.filename)) == current_script_dir:
            return frame

    # No match so return the topmost frame
    return frame


def format_landing_excel_receipt(ticket, flights, landings, fees, contract_number, landing_fee):
    # For each flight, get the ordinal for each landing type (e.g., dropoff 1, dropoff 2, pickup 1, pickup 2)
    these_landings = landings.loc[landings.flight_id.isin(flights.id)]
    for flight_id, flight in these_landings.groupby(['flight_id', 'landing_type']):
        these_landings.loc[flight.index, 'landing_order'] = list(range(1, len(flight) + 1))
    these_landings['landing_type_order'] = these_landings.landing_type + '_' + these_landings.landing_order.astype(
        int).astype(str)

    # Pivot the table for passengers and landings so that each row represents 1 flight
    pivoted_passengers = these_landings.pivot(index='flight_id', columns='landing_type_order', values='n_passengers')
    pivoted_locations = these_landings.pivot(index='flight_id', columns='landing_type_order', values='location')
    pivoted_passengers.rename(columns={c: c + '_passengers' for c in pivoted_locations.columns}, inplace=True)
    pivoted_locations.rename(columns={c: c + '_location' for c in pivoted_locations.columns}, inplace=True)

    receipt = flights.drop(columns='flight_id') \
        .merge(pivoted_passengers, left_on='id', right_index=True, how='left') \
        .merge(pivoted_locations, left_on='id', right_index=True, how='left') \
        .merge(fees.fillna(0), left_on='id', right_on='flight_id',
               how='left')  # left join in case there aren't any fees
    receipt['departure_date'] = receipt.departure_datetime.dt.strftime('%m/%d/%Y')
    receipt['departure_time'] = receipt.departure_datetime.dt.strftime('%I:%M %p')

    # Concat all notes
    these_landings['landing_notes'] = these_landings.drop_duplicates(['flight_id', 'location']) \
                                          .loc[:, ['justification', 'notes']] \
        .apply(lambda x: ('; '.join(x.fillna('').astype(str)).strip('; ')), axis=1) \
        .fillna('')
    notes = these_landings.groupby('flight_id').apply(lambda group: '; '.join(group.landing_notes.dropna()).strip('; '))
    notes.index = notes.index.tolist()
    receipt['landing_notes'] = receipt.merge(pd.DataFrame({'flight_landing_notes': notes, 'flight_id': notes.index}),
                                             on='flight_id').flight_landing_notes
    receipt['all_notes'] = receipt.loc[:, [c for c in ['landing_notes', 'notes'] if c in receipt]] \
        .apply(lambda x: ('; '.join(x.fillna('').astype(str)).strip('; ')), axis=1) \
        .fillna('')
    # receipt['n_fee_pax'] = receipt.n_passengers
    receipt['n_passengers'] = receipt[
        [c for c in receipt if c.endswith('_passengers') and c not in ('n_passengers', 'n_fee_passengers')]].sum(axis=1)
    # receipt['n_fee_passengers'] = receipt.n_fee_pax

    # Fill any numeric columns with 0. All of the columns that don't yet exist in the receipt (e.g., n_pickup_3)
    #   will still be blank in the final receipt
    numeric_columns = [c for c in receipt if c.startswith('n_')] + ['fee']
    receipt[numeric_columns] = receipt[numeric_columns].fillna(0)

    receipt_info = pd.DataFrame([{'operator': receipt.operator_code.iloc[0],
                                  'ticket': ticket,
                                  'date': datetime.now().strftime('%m/%d/%Y'),
                                  'pax_fee': landing_fee,
                                  'contract': contract_number}]) \
        .reindex(columns=['ticket', 'date', 'operator', 'contract', 'pax_fee'])

    return receipt.reindex(columns=LANDING_RECEIPT_COLUMNS), receipt_info


def format_landing_html_receipt(receipt):
    '''
    Format a dataframe to include in an email as an HTML table. Mostly just reduce the number of columns.
    :param receipt:
    :return:
    '''

    _html_receipt_columns = ['Date', 'Time', 'Tail Number', 'Aircraft Type', 'Scenic Route', 'Landing Locations',
                             'Total Pax', 'Fee']
    # Combine all location columns
    receipt['Landing Locations'] = [', '.join(row.dropna().unique()) for _, row in
                                    receipt.loc[:, [c for c in receipt.columns if c.endswith('location')]].iterrows()]

    # If no scenic routes were given, this is probably not an operator that uses them so drop the column
    if receipt.scenic_route.any():
        receipt['Scenic Route'] = receipt.scenic_route
    receipt['Date'] = receipt.departure_datetime.dt.strftime('%m/%d/%Y')
    receipt['Time'] = receipt.departure_datetime.dt.strftime('%I:%M %p')
    receipt['Fee'] = receipt.fee.apply('${:.2f}'.format)
    receipt['Total Pax'] = receipt.n_passengers.astype(int)
    receipt = receipt \
        .rename(columns={'registration': 'Tail Number', 'aircraft_type': 'Aircraft Type'}) \
        .reindex(columns=_html_receipt_columns).fillna('')

    return receipt


def save_excel_receipt(template_path, receipt_dir, ticket, data_type, data, header_path, sheet_password,
                       info=pd.DataFrame()):
    ''' Save a formatted data frame as an Excel file to attach to notification emails to submitters'''

    # Remove the "data" sheet from the template so it can be replaced with the dataframe
    xl_path = os.path.join(receipt_dir, '{}_receipt_ticket{}.xlsx'.format(data_type, ticket))
    wb = openpyxl.load_workbook(template_path)
    wb.remove(wb['data'])
    if 'info' in wb.sheetnames and len(info):
        wb.remove(wb['info'])

    # Add the header img
    img = openpyxl.drawing.image.Image(header_path)
    img.anchor = 'E1'
    ws = wb['submission_receipt']
    ws.add_image(img)
    wb.save(xl_path)
    wb.close()

    # Add the data and info if it was given. Use pandas instead of the openpyxl.Workbook object because writing a df to
    #   Excel with openpyxl just overwrites the existing workbook
    with pd.ExcelWriter(xl_path, engine='openpyxl', mode='a') as writer:
        data.to_excel(writer, sheet_name='data', index=False)
        if len(info):
            info.to_excel(writer, sheet_name='info', index=False)

    # Protect each sheet from edits. Do this after writing the data to new sheets
    wb = openpyxl.load_workbook(xl_path)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        ws.protection.password = sheet_password
        ws.protection.sheet = True
        ws.protection.enable()
        if sheet_name != 'submission_receipt':
            ws.sheet_state = 'hidden'
    wb.save(xl_path)
    wb.close()

    return xl_path


def validate_excel_landings(df, engine, row_offset=9, error_handling='raise'):
    excel_errors = []
    excel_warnings = []

    if df.departure_datetime.isnull().any():
        excel_errors.append('There are flights without departure dates/times or with dates/times in invalid formats')
        return excel_errors, excel_warnings

    df['row_str_id'] = df.departure_datetime.dt.strftime('%m/%d/%y %I:%M') + \
                       pd.Series(df.index + row_offset).apply(lambda x: ' (row %s)' % x)

    for i, row in df.drop('n_passengers', axis=1).iterrows():
        location_cols = row[df.columns[df.columns.str.contains(r'\d_location')]].dropna().index
        passenger_cols = row[df.columns[df.columns.str.contains(r'\d_passengers')]].loc[~row.isna() & (row.astype(str) != '')].index

        # Check that all flights have at least one landing
        n_locations = len(location_cols)
        n_passengers = len(passenger_cols)
        if not n_locations + n_passengers:
            excel_warnings.append('No landings entered for the flight departing at {row_str_id}'.format(**row))
            continue
        # If so, check that at least one landing location/passenger count has been recorded per row
        else:
            if not n_locations:
                excel_errors.append('No landing location given for the flight departing at {row_str_id}'.format(**row))
                continue
            if not n_passengers:
                excel_errors.append('No passenger count given for the flight departing at {row_str_id}'.format(**row))
                continue

        # Check that for each passenger count there's a location and vice versa
        missing_locations = passenger_cols[
            ~passenger_cols.isin([c.replace('_location', '_passengers') for c in location_cols])]
        missing_passengers = location_cols[
            ~location_cols.isin([c.replace('_passengers', '_location') for c in passenger_cols])]
        if len(missing_locations):
            excel_errors.append(
                'For the flight departing %s, the landings columns are missing a location: %s' % (
                row.row_str_id, ', '.join(missing_locations.str.replace('passengers', 'location'))))

        if len(missing_passengers):
            excel_errors.append(
                'For the flight departing %s, the landings at the following locations are missing passenger counts: %s' % (
                row.row_str_id, ', '.join(missing_passengers)))

    # Check that there's a justification given for any scenic landings that require it
    scenic_landings = df.dropna(subset=['scenic_1_location'])
    sql = 'SELECT name, scenic_data_entry_note FROM landing_locations WHERE scenic_data_entry_note IS NOT NULL'
    needs_justification = pd.read_sql(sql, engine)
    without_justification = scenic_landings.loc[
        scenic_landings.scenic_1_location.isin(needs_justification['name']) &
        scenic_landings.notes.isnull()
        ]
    if len(without_justification):
        excel_warnings.append('Flights that departed at the following times need justification for scenic landing'
                        ' locations: %s' % ', '.join(without_justification.row_str_id))

    taxi_landing_fields = df.loc[:,
                          df.columns[df.columns.str.contains(r'\d_location') & ~df.columns.str.contains('scenic')]]
    taxi_flights_without_location = df.loc[
        (taxi_landing_fields == 'Other').any(axis=1) &
        (df.notes.isna() | (df.notes.astype(str).str.strip() == ''))
        ]
    if len(taxi_flights_without_location):
        excel_warnings.append('Flights that departed at the following times need a location specified in the notes for 1 or'
                      ' more taxi dropoffs/pickups locations: %s' % ', '.join(without_justification.row_str_id))

    # If errors should be raised, raise for either errors or warnings
    # Otherwise, errors and warnings will be handled differently
    if len(excel_errors) and error_handling == 'raise':
        raise ValueError('Invalid or missing values found in file:\n\t-%s' % '\n\t-'.join(excel_errors + excel_warnings))

    return excel_errors, excel_warnings


def process_excel_landings(excel_path, landings_conn, submitted_info, data_steward,
                           flights_agol_columns, error_handling=None, row_offset=9, data_sheet_name='data', info_sheet_name='info'):

    # Read data from the downloaded Excel file
    #   Get the column names from the 'data' sheet
    columns = ['departure_date', 'departure_time'] + pd.read_excel(excel_path, 'data').columns.tolist()[1:]
    data = pd.read_excel(excel_path, 'landings', skiprows=row_offset - 1, header=None, names=columns, usecols=range(len(columns)))\
        .dropna(subset=['departure_date', 'departure_time', 'registration'])
    info = pd.read_excel(excel_path, info_sheet_name).squeeze()
    ticket = submitted_info.ticket
    agol_id = submitted_info.parentglobalid

    # Combine date and time
    data.departure_time = pd.to_datetime(data.departure_time.astype(str)).dt.time # make sure all departure_times are times and not datetimes
    data['departure_datetime'] = data.apply(lambda row: pd.datetime.combine(row.departure_date, row.departure_time), axis=1)
    data.drop(columns=['departure_date', 'departure_time'], inplace=True)

    excel_errors, excel_warnings = validate_excel_landings(data, landings_conn, row_offset, error_handling=error_handling)
    # If there were any errors, the data shouldn't be processed. Add an error to the notifications email and return
    #   empty dataframes
    if len(excel_errors):
        html_li = (
            '<li>The Excel file {excel_file} could not be processed because it produced the following errors:'
            '<ul><li>{error_str}</li></ul>').format(excel_file=os.path.basename(excel_path), error_str='</li><li>'.join(excel_errors))
        MESSAGES.append(
            {'ticket': ticket, 'message': html_li, 'recipients': data_steward, 'type': 'landings', 'level': 'error', 'attachment': excel_path}
        )
        # Return dummy data
        return  pd.DataFrame(flights_agol_columns), pd.DataFrame(columns=AGOL_LANDING_COLUMNS)

    # Warnings are fine, but the data steward(s) needs to be notified
    if len(excel_warnings):

        html_li = (
            '<li>The Excel file {excel_file} was processed, but it produced the following warnings:'
            '<ul><li>{warning_str}</li></ul>'
        ).format(excel_file=os.path.basename(excel_path), warning_str='</li><li>'.join(excel_warnings))
        MESSAGES.append(
            {'ticket': ticket, 'message': html_li, 'recipients': data_steward, 'type': 'landings', 'level': 'warning', 'attachment': excel_path}
        )

    data['flight_id'] = data.registration + '_' + \
                        data.departure_datetime.dt.floor('15min').dt.strftime('%Y%m%d%H%M')
    data['operator'] = info.operator_code
    data['globalid'] = [str(uuid.uuid4()) for _ in range(len(data))]#agol_id + '-' + data.index.astype(str)
    data['submission_type'] = 'landings'

    landings = data.drop(columns=['n_passengers', 'n_fee_passengers'])\
        .melt(
            id_vars=[c for c in data if not c.endswith('_passengers')],
            value_name='n_passengers'
        ).dropna(subset=['n_passengers'])
    landings['landing_type'] = landings.variable.apply(lambda x: x.split('_')[0])
    landings['location'] = [landings.loc[i, c.replace('_passengers', '_location')] for i, c in landings.variable.iteritems()]
    landings['n_passengers_scenic'] = landings.loc[landings.landing_type == 'scenic', 'n_passengers']
    landings['n_passengers_dropoff'] = landings.loc[landings.landing_type == 'dropoff', 'n_passengers']
    landings['n_passengers_pickup'] = landings.loc[landings.landing_type == 'pickup', 'n_passengers']

    # Group landings into individual landings (e.g., a pickup and dropoff occurred at the same time)
    grouped_landings = landings.groupby(['flight_id','location'])
    landing_types = grouped_landings.landing_type.apply(lambda x: ','.join(x))
    grouped_landings = grouped_landings.sum()
    grouped_landings['landing_type'] = landing_types
    grouped_landings = grouped_landings.reset_index()
    grouped_landings['parentglobalid'] = grouped_landings\
        .merge(landings.drop_duplicates(subset=['flight_id']), on='flight_id').globalid

    landing_locations = db_utils.get_lookup_table(table='landing_locations', index_col='name', value_col='code',
                                               conn=landings_conn)
    landings = grouped_landings.rename(columns={v: k for k, v in LANDINGS_COLUMNS.items()})\
        .reindex(columns=AGOL_LANDING_COLUMNS)\
        .replace({'landing_location': landing_locations}) # data come as names, not codes
    landings.globalid = [str(uuid.uuid4()) for _ in range(len(landings))]

    aircraft_types = db_utils.get_lookup_table(table='aircraft_types', index_col='name', value_col='code',
                                               conn=landings_conn)
    flights = data.rename(columns={v: k for k, v in LANDINGS_FLIGHT_COLUMNS.items()})\
        .reindex(columns=flights_agol_columns)\
        .replace({'landing_aircraft_type': aircraft_types}) # data come as names, not codes
    flights.is_from_excel = True
    flights.ticket = ticket
    flights.submission_time = submitted_info.submission_time

    return flights, landings


def process_excel_submission(excel_path, landings_conn, flight_info, param_dict, flight_columns, error_handling=None):
    '''
    Helper function to process a single row of the flight table that's an Excel landing submission. The main reason
    for the helper is to wrap process_excel_landings() to conveniently call it from multiple places

    :param excel_path: path to the downloaded Excel (or zipped Excel) file
    :param landings_conn: SQLAlchemy DBAPI connection
    :param flight_info: Pandas.Series of the flight table row
    :param param_dict: parameters from poll_feature_service_params.json
    :param flight_columns: all columns for the flight table
    :param error_handling: flag to set the error logging level
    :return: whatever process_excel_landings() returns
    '''
    if excel_path.endswith('.zip'):
        zip_path = excel_path
        with zipfile.ZipFile(excel_path) as z:
            # Zipping the Excel file is only necessary to get beyond Survey123's file type filter so just
            #   process the first file because there shouldn't be more than one
            excel_path = z.extract(z.namelist()[0], path=os.path.dirname(excel_path))
        os.remove(zip_path)

    result = process_excel_landings(
        excel_path,
        landings_conn,
        flight_info,
        param_dict['landing_data_stewards'],
        flight_columns,
        error_handling=error_handling,
        **param_dict['excel_landing_template']
    )
    os.remove(excel_path)

    return result


def import_landings(flights, ticket, landings_conn, sqlite_path, landings, receipt_dir, receipt_template,
                    receipt_header, sheet_password, data_steward):
    '''
    Process and import landing data into Postgres backend. Send receipt to the submitter

    :param flights: dataframe of newly submitted flights
    :param landings_conn: connection object to Postgres DB of landings
    :param sqlite_path: downloaded data from AGOL
    :param landings: data frame of landings
    :param receipt_dir: directory to save Excel file submission receipts
    :param receipt_template: Excel file to use as template for making receipts
    :return:
    '''

    _html_receipt_columns = ['departure_datetime', 'scenic_route', 'registration', 'aircraft_type', 'submission_time']

    # Create flight IDs
    landing_flights = flights.loc[flights.submission_type == 'landings'] \
        .dropna(axis=1, how='all')  # columns for all track submissions will be empty so drop them
    landing_flights.landing_tail_number = landing_flights.landing_tail_number.str.upper()
    landing_flights.rename(columns=LANDINGS_FLIGHT_COLUMNS, inplace=True)
    landing_flights['flight_id'] = landing_flights.registration + '_' + \
                                   landing_flights.departure_datetime.dt.floor('15min').dt.strftime('%Y%m%d%H%M')
    existing_flight_ids = pd.read_sql("SELECT flight_id FROM flights", landings_conn).squeeze(axis=1)
    new_flights = landing_flights.loc[~landing_flights.flight_id.isin(existing_flight_ids) &
                                      ~landing_flights.departure_datetime.isnull()]

    # Check for flights with no landings
    landings_per_flight = landing_flights \
        .loc[~landing_flights.is_from_excel]\
        .merge(landings, left_on='agol_global_id', right_on='parentglobalid', how='left') \
        .groupby('flight_id') \
        .parentglobalid \
        .count()
    no_landings = landings_per_flight[landings_per_flight == 0]
    if len(no_landings):
        flights_without_landings = landing_flights.loc[landing_flights.flight_id.isin(no_landings.index)] \
            .reindex(columns=_html_receipt_columns + ['objectid'])
        html_li = (
            '<li>There were no landings submitted for the following flights:<br>{table}<br></li>') \
            .format(table=get_html_table(flights_without_landings))
        MESSAGES.append(
            {'ticket': ticket, 'message': html_li, 'recipients': data_steward, 'type': 'landings', 'level': 'warning'})
        # Remove the flights without landings so they don't get imported
        new_flights = new_flights.loc[~new_flights.flight_id.isin(no_landings.index)]

    # Check if there are any new flights. If not, warn the submitter(s) that no new flights were found
    if len(new_flights) == 0:
        html_li = (
            '<li>All landings submitted with this ticket were already reported (according to the tail number and departure time). Flights submitted with this ticket:<br>{table}<br></li>').format(
            table=get_html_table(landing_flights.reindex(columns=_html_receipt_columns + ['objectid'])))
        MESSAGES.append(
            {'ticket': ticket, 'message': html_li, 'recipients': data_steward, 'type': 'landings', 'level': 'warning'}
        )
        return None, None, None

    aircraft_types = db_utils.get_lookup_table(table='aircraft_types', conn=landings_conn)
    operators = db_utils.get_lookup_table(table='operators', conn=landings_conn)
    landing_locations = db_utils.get_lookup_table(table='landing_locations', conn=landings_conn)
    fee_per_passenger = db_utils.get_lookup_table(table='operators', index_col='code', value_col='fee_per_passenger',
                                          conn=landings_conn)[landing_flights.operator_code.iloc[0]]

    if len(new_flights) != len(landing_flights):
        # If some flights were duplicates, warn the landing data steward(s)
        operator = operators[landing_flights.operator_code.iloc[0]]
        duplicate_flights = landing_flights.loc[~landing_flights.flight_id.isin(new_flights.flight_id)] \
            .reindex(columns=_html_receipt_columns + ['objectid'])
        html_li = (
            '<li>Some landings were already reported (according to the tail number and departure time). You might'
            ' want to warn {operator} that landings for the following duplicated flights were already submitted:'
            '<br>{table}<br></li>').format(operator=operator, table=get_html_table(duplicate_flights))
        MESSAGES.append(
            {'ticket': ticket, 'message': html_li, 'recipients': data_steward, 'type': 'landings', 'level': 'warning'})

    new_flights['submission_method'] = 'survey123'
    new_flights['source_file'] = sqlite_path
    new_flights['fee_per_passenger'] = fee_per_passenger
    new_flights.reindex(columns=LANDINGS_FLIGHT_COLUMNS.values()) \
        .to_sql('flights', landings_conn, if_exists='append', index=False)
    global_ids = pd.read_sql("SELECT id, agol_global_id FROM flights WHERE flight_id IN ('%s')"
                             % "', '".join(new_flights.flight_id),
                             landings_conn)

    # Calculate concessions fees and attach numeric IDs to landings and fees
    # some landings might have multiple landing types selected so split them
    landings['sort_order'] = range(len(landings))
    passenger_cols = pd.Series([c for c in landings.columns if c.startswith('n_passengers')])

    def clean_n_passenger_cols(row, cols):
        row.loc[cols.loc[~cols.str.endswith(row.landing_type)]] = np.nan
        return row

    landings['landing_type'] = landings.landing_type.str.split(',')
    landings = pd.DataFrame([clean_n_passenger_cols(row, passenger_cols) for i, row in
                             landings.explode('landing_type').iterrows()]).reset_index()
    landings['index'] = landings.index
    landings_with_fees = pd.read_sql("SELECT name FROM landing_types WHERE fee_charged;", landings_conn).squeeze(
        axis=1).tolist()

    passengers = landings \
        .melt(id_vars='globalid',
              value_vars=passenger_cols,
              value_name='n_passengers',
              var_name='landing_type') \
        .dropna(subset=['n_passengers'])  # '''
    passengers['landing_type'] = passengers.landing_type.apply(lambda x: x.split('_')[-1])

    # Get the flight ID for the landings table
    landings = landings.merge(global_ids.rename(columns={'id': 'flight_id'}), left_on='parentglobalid',
                              right_on='agol_global_id')

    landings = passengers.merge(landings.drop(columns='landing_type'), on='globalid') \
        .drop_duplicates(subset=['globalid', 'landing_type']) \
        .rename(columns=LANDINGS_COLUMNS) \
        .reindex(columns=LANDINGS_COLUMNS.values()) \
        .dropna(subset=['flight_id']) \
        .sort_values('sort_order')

    # INSERT into backend
    landings.drop(columns='sort_order') \
        .to_sql('landings', landings_conn, if_exists='append', index=False)  # '''

    global DATA_PROCESSED
    DATA_PROCESSED = True

    contracts = db_utils.get_lookup_table(table='operators', index_col='code', value_col='contract_number',
                                          conn=landings_conn)
    this_contract = contracts[new_flights.operator_code.iloc[0]]

    # Read from view that calculates fees automatically
    fees = pd.read_sql(
        "SELECT * FROM concession_fees_view WHERE flight_id IN (%s)" % ','.join(global_ids.id.astype(str)),
        landings_conn)

    # Format the table for the receipt(s)
    # Replace codes with names
    new_flights.reset_index(inplace=True)
    new_flights['id'] = new_flights.merge(global_ids.reset_index(), on='agol_global_id').id
    # INR name and code are the same so don't replace it because pandas throws an error
    landings.replace({'location': {k: v for k, v in landing_locations.items() if k != v}},
                     inplace=True)  # {c: landing_locations for c in receipt.columns if c.endswith('location')}, inplace=True)

    new_flights.replace({'aircraft_type': aircraft_types, 'operator_code': operators}, inplace=True)

    ticket = new_flights.ticket.iloc[0]
    # for ticket, flights in new_flights.groupby('ticket'):
    receipt, info = format_landing_excel_receipt(ticket, new_flights, landings, fees, this_contract, fee_per_passenger)
    receipt_path = save_excel_receipt(receipt_template, receipt_dir, ticket, 'landing_data', receipt, receipt_header,
                                      sheet_password, info)
    html_receipt = format_landing_html_receipt(receipt)

    return receipt_path, html_receipt, new_flights


def get_unique_filename(path, filename_suffix=''):
    _, ext = os.path.splitext(path)
    ext = filename_suffix + ext
    c = 1
    new_path = path
    while os.path.isfile(new_path):
        previous_count = c - 1 if c > 1 else c
        new_path = path.replace(ext, '') + '_{c}{ext}'.format(c=c, ext=ext)
        c += 1

    return new_path


def prepare_track(track_path, attachment_info, import_params, submissions, submitter, operator_codes, mission_codes,
                  attachment_dir, data_steward, web_data_dir, db_columns):
    '''Prepare the track file to be read by the editing web app (inpdenards.nps.doi.net/track-editor.html). It needs to
    be in format that both the web app and the rest of import_track functions will understand'''

    fname = os.path.basename(track_path)

    def delete_attachment():
        basename, _ = os.path.splitext(track_path)
        for path in glob(basename + '*'):
            try:
                os.remove(path)
            except:
                pass

    # Get submission info
    this_agol_id = attachment_info['REL_GLOBALID'].replace('{', '').replace('}', '')
    this_submission = submissions.loc[submissions.globalid == this_agol_id].squeeze()
    # This file should have been deleted when the AGOL record was
    if len(this_submission) == 0:
        # delete the track and attachment info json and return
        delete_attachment()
        return

    for k, v in this_submission.iteritems():
        # most object types can't be serialized as JSON, so make them strs
        attachment_info[k] = str(v) if v and v != 0 else ''
    attachment_info['submitter'] = submitter
    attachment_info['submitter_notes'] = attachment_info['tracks_notes']

    try:
        gdf = import_track.format_track(track_path,
                                        registration=attachment_info['tracks_tail_number'],
                                        submission_method='survey123',
                                        **import_params)
    except Exception as e:
        tb_frame = get_traceback_frame()
        html_li = (
            '<li>An error occurred while processing the file {track_path} on line {lineno} of {script}: {error}. AGOL flight table Object ID: {agol_id}</li>') \
            .format(track_path=track_path, lineno=tb_frame.lineno, script=os.path.basename(tb_frame.filename), error=e,
                    agol_id=attachment_info['objectid'])  # , submitter_clause=submitter_error_clause)
        # Add a message per recipient so messages can be easily parsed per recipient later
        MESSAGES.append(
            {'ticket': attachment_info['ticket'], 'message': html_li, 'recipients': data_steward, 'type': 'tracks',
             'level': 'error'})
        ERRORS.append(traceback.format_exc())
        return

    if not len(gdf):
        html_li = (
            '<li>The file {track_path} did not contain any readable tracks. AGOL flight table Object ID: {agol_id}</li>').format(
            track_path=track_path, agol_id=attachment_info['objectid'])
        MESSAGES.append(
            {'ticket': attachment_info['ticket'], 'message': html_li, 'recipients': data_steward, 'type': 'tracks',
             'level': 'error'})
        return

    # Drop any columns that aren't in the flights or flight_points tables
    gdf.drop(columns=gdf.columns[~gdf.columns.isin(db_columns + ['segment_id', 'geometry'])], inplace=True)

    attachment_info['operator_code'] = this_submission.tracks_operator

    if attachment_info['tracks_mission'] in mission_codes:
        attachment_info['nps_mission_code'] = attachment_info[
            'tracks_mission']  # , mission_codes[attachment_info['tracks_mission']]
    else:
        attachment_info['nps_mission_code'] = ''

    attachment_info['nps_work_group'] = attachment_info['tracks_work_group']
    attachment_info['nps_work_group_other'] = attachment_info['tracks_other_work_group']

    attachment_info['source_file'] = get_unique_filename(os.path.join(import_track.ARCHIVE_DIR, fname))

    # Dump each line segment to a geojson string. Write a list of these strings as a JSON file to be read
    #   by the web app
    geojson_strs = track_to_json(gdf, attachment_info)
    basename, ext = os.path.splitext(fname)
    basename = re.sub('\W', '_', basename)  # strip chars for Javascript (i.e., escape and CSS selectors)
    json_path = os.path.join(web_data_dir, basename) + '_geojsons.json'
    # make sure an existing file doesn't get overwritten if it has same name
    # if os.path.isfile(json_path):
    json_path = get_unique_filename(json_path, filename_suffix='_geojsons')

    with open(json_path, 'w') as j:
        json.dump(geojson_strs, j)

    global DATA_PROCESSED
    DATA_PROCESSED = True

    # Copy to archive dir
    try:
        shutil.copy(track_path, attachment_info['source_file'])
        # Delete the file and attachment info json since the data were processed successfully
        # This won't throw an error, but it should only be executed if the above line succeeds
        delete_attachment()
    except:
        html_li = (
            '<li>Unable to copy {track_path} to the track archive directory {archive_dir}. This file should be copied manually.</li>').format(
            track_path=track_path, archive_dir=import_track.ARCHIVE_DIR)
        MESSAGES.append(
            {'ticket': attachment_info['ticket'], 'message': html_li, 'recipients': data_steward, 'type': 'tracks',
             'level': 'warning'})

    return json_path


def prepare_track_data(param_dict, download_dir, tracks_conn, submissions):

    import_params = param_dict['import_params'] if 'import_params' in param_dict else {}
    attachment_dir = os.path.join(download_dir, 'attachments')
    operator_codes = db_utils.get_lookup_table(table='operators', conn=tracks_conn)
    mission_codes = db_utils.get_lookup_table(table='nps_mission_codes', conn=tracks_conn)
    data_steward = param_dict['track_data_stewards']

    db_columns = pd.concat([db_utils.get_db_columns('flights', conn=tracks_conn),
                            db_utils.get_db_columns('flight_points', conn=tracks_conn)]) \
        .drop_duplicates() \
        .to_list()

    # If for some reason the dir got removed, create it
    if not os.path.isdir(attachment_dir):
        os.mkdir(attachment_dir)

    submitted_files = {}
    geojson_paths = []

    # Unzip all zipped files and add the extracted files to the list of tracks to process
    processed_unzipped_files = [] # keep track of processed files from zip files so they can be skipped in next for loop
    for zip_path in glob(os.path.join(attachment_dir, '*.zip')):
        try:
            attachment_info = get_attachment_info(zip_path, submissions)
        except Exception as e:
            html_li = ('<li>An error occurred while trying to read metadata for {file}: {error}.</li>').format(
                file=zip_path, error=e)
            MESSAGES.append(
                {'ticket': -1, 'message': html_li, 'recipients': data_steward, 'type': 'tracks',
                 'level': 'error'})
            ERRORS.append(traceback.format_exc())
            continue
        submitter = get_submitter_email(attachment_info['Creator'], param_dict['agol_users'])
        unzipped_files = []
        with zipfile.ZipFile(zip_path) as z:
            for fname in z.namelist():
                try:
                    z.extract(fname, attachment_dir)
                except Exception as e:
                    html_li = (
                        '<li>An error occurred while trying to extract {file} from {zip}: {error}. AGOL flight table Object ID: {agol_id}.</li>').format(
                        file=fname, zip=zip_path, error=e, agol_id=attachment_info['objectid'])
                    MESSAGES.append(
                        {'ticket': attachment_info['ticket'], 'message': html_li, 'recipients': data_steward,
                         'type': 'tracks', 'level': 'error'})
                    ERRORS.append(traceback.format_exc())
                unzipped_files.append(fname)

        for fname in unzipped_files:
            basename, ext = os.path.splitext(fname)
            if ext not in import_track.READ_FUNCTIONS:
                MESSAGES.append(
                    {'ticket': attachment_info['ticket'], 'message': '%s is not in an accepted file format' % fname,
                     'recipients': data_steward, 'type': 'tracks', 'level': 'error'})
                continue
            path = os.path.join(attachment_dir, fname)
            try:
                geojson_path = prepare_track(path, attachment_info, import_params, submissions, submitter,
                                             operator_codes, mission_codes, attachment_dir, data_steward,
                                             param_dict['web_data_dir'], db_columns)
                geojson_paths.append(geojson_path)
                submitted_files[attachment_info['REL_GLOBALID'].strip('{}')] = os.path.basename(zip_path)

            except Exception as e:
                html_li = (
                    '<li>An error occurred while trying to process {file} from {zip}: {error}. AGOL flight table Object ID: {agol_id}.</li>').format(
                    file=fname, zip=zip_path, error=e, agol_id=attachment_info['objectid'])
                MESSAGES.append({'ticket': attachment_info['ticket'], 'message': html_li, 'recipients': data_steward,
                                 'type': 'tracks', 'level': 'error'})
                ERRORS.append(traceback.format_exc())
            processed_unzipped_files.append(fname)  # add to list of processed files so the next for loop will skip them

    # Find all the files that were downloaded without being zipped. This will only include files with valid extensions
    #   because Survey123 will reject all others
    for ext in import_track.READ_FUNCTIONS:
        for path in glob(os.path.join(attachment_dir, '*%s' % ext)):
            fname = os.path.basename(path)

            if fname in submitted_files.values() or fname in processed_unzipped_files:
                continue

            attachment_info = get_attachment_info(path, submissions, ext)

            submitter = get_submitter_email(attachment_info['Creator'], param_dict['agol_users'])
            try:
                geojson_path = prepare_track(path, attachment_info, import_params, submissions, submitter,
                                             operator_codes, mission_codes, attachment_dir, data_steward,
                                             param_dict['web_data_dir'], db_columns)
                geojson_paths.append(geojson_path)
                submitted_files[attachment_info['REL_GLOBALID'].strip('{}')] = os.path.basename(path)
            except Exception as e:
                html_li = ('<li>An error occurred while trying to process {file}: {error}.</li>').format(file=path,
                                                                                                         error=e)
                MESSAGES.append({'ticket': attachment_info['ticket'], 'message': html_li, 'recipients': data_steward,
                                 'type': 'tracks', 'level': 'error'})
                ERRORS.append(traceback.format_exc())

    return geojson_paths, pd.Series(submitted_files)


def get_track_receipt_email(submission, ticket, sender, agol_users):
    _html_track_columns = ['Submission Time', 'Tail Number', 'NPS Mission Code', 'NPS Work Group', 'Filename', 'Notes']

    recipient = get_submitter_email(submission.submitter.iloc[0], agol_users)
    if not recipient:
        return {}
    start_time = submission.submission_time.min()
    end_time = submission.submission_time.max()

    submission = submission \
        .rename(columns={
        'tracks_tail_number': 'Tail Number',
        'tracks_mission': 'NPS Mission Code',
        'tracks_work_group': 'NPS Work Group',
        'tracks_notes': 'Notes',
        'submission_time': 'Submission Time'}
    ) \
        .reset_index() \
        .reindex(columns=_html_track_columns)

    # If this was submitted by someone other than an NPS employee, drop the NPS columns
    nps_columns = pd.Series([c for c in submission if c.startswith('NPS')])
    # null_nps_columns = nps_columns[submission.loc[:, nps_columns].isnull().all(axis=0).values]
    if not recipient.endswith('@nps.gov'):  # len(null_nps_columns):
        submission.drop(columns=nps_columns, inplace=True)

    html = get_email_html(submission.fillna(''), start_time, end_time,
                          column_widths={'Notes': '50px; max-width: 100px'})

    message = {'subject': 'Flight track submission receipt - ticket #%s' % ticket,
               'message_body': html,
               'recipients': [recipient],
               'sender': sender,
               'message_body_type': 'html'}

    return message


def compose_error_notifications(tickets, param_dict):
    messages = pd.DataFrame(MESSAGES)
    messages['recipient_str'] = messages.recipients.apply(str)
    html_emails = []
    for (ticket, _, submission_type), info in messages.groupby(['ticket', 'recipient_str', 'type']):
        ul_template = r'''
        <h4>{type}:</h4>
        <ul>
            {li_elements}
        </ul>
        <br>
        '''
        warnings = info.loc[info.level == 'warning', 'message']
        errors = info.loc[info.level == 'error', 'message']
        warning_html = ul_template.format(type='Warnings', li_elements=''.join(warnings))
        error_html = ul_template.format(type='Errors', li_elements=''.join(errors))
        attachments = info.attachment.unique().tolist() if 'attachment' in info.columns else []

        if ticket == -1:
            message = '''
            <html>
                <head></head>
                <body>
                    <p>Flight data were recently submitted that resulted in warnings and/or errors.
                        <br>
                        {warnings}
                        {errors}
                    </p>
                    You can inspect these files at {attachments_dir}.
                    <br>
                </body>
            </html>
            '''.format(
                warnings=warning_html if len(warnings) else '',
                errors=error_html if len(errors) else '',
                attachments_dir=os.path.join(param_dict['download_dir'], 'attachments')
            )
            subject = 'Issues with recently submitted {} data files'.format(submission_type)
        else:
            submission_info = tickets.loc[
                (tickets.ticket == int(ticket)) & (tickets.submission_type == submission_type)].squeeze()
            email_address = get_submitter_email(submission_info.submitter, param_dict['agol_users']).split('<')[-1].strip('>')
            # if submission_type == 'tracks':
            concluding_message = r'<br><br><span>If you cannot resolve the issue(s), try contacting the submitter directly if they are an NPS employee. If the submitter is a commercial flight operator, notify Commercial Services of the problem and they will contact the submitter. </span>'
            # else:
            #    concluding_message = ''
            message = '''
            <html>
                <head></head>
                <body>
                    <p>Flight data were recently submitted by the AGOL user {submission_info.submitter} that resulted in warnings and/or errors.
                        <br>
                        {warnings}
                        {errors}
                        <span><strong>Submitter email:</strong> {email_address}</span>
                        <br>
                        <span><strong>Submission time:</strong> {submission_info.submission_time}</span>
                        {concluding_message}
                    </p>
                </body>
            </html>
            '''.format(submission_info=submission_info, email_address=email_address,
                       warnings=warning_html if len(warnings) else '',
                       errors=error_html if len(errors) else '',
                       concluding_message=concluding_message if submission_type == 'tracks' else '',
                       delete_url=param_dict['delete_features_url'])
            subject = 'Issues with recent {submission_type} data submssion - ticket #{ticket:.0f}' \
                .format(submission_type=submission_type, ticket=int(ticket))
        html_emails.append(dict(subject=subject,
                                message_body=message,
                                recipients=info.recipients.iloc[0],
                                sender=param_dict['mail_sender'],
                                message_body_type='html',
                                attachments=attachments
                                )
                           )

    return html_emails


def poll_feature_service(log_dir, download_dir, param_dict, ssl_cert, landings_conn, tracks_conn):
    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)

    # Get the last time any data were downloaded
    script_name = os.path.basename(__file__).rstrip('.py')
    log_info = read_logs(log_dir, script_name)

    # Set default time. If this is never reset, there are no logs because this is the first time the script is being
    #   run (or the logs got deleted), so set the last_download_time to something that will return all records
    last_download_time = '1970-1-1 09:00:00'
    if len(log_info):
        download_info = log_info.loc[log_info.data_downloaded]
        if len(download_info):  # there has been at least one download in the last LOG_CACHE_DAYS
            last_download_time = download_info.timestamp.max()

    # Get last submission time
    agol_credentials = param_dict['agol_credentials']
    service_url = agol_credentials['service_url']
    try:
        token = download.get_token(**agol_credentials, ssl_cert=ssl_cert)
        service_info = download.get_service_info(service_url, token, ssl_cert)
    except:
        EMAILS.append({'message_body': 'An unexpected error occurred while generating a REST API token:\n%s'
                                       % traceback.format_exc(),
                       'subject': 'Error while running %s' % __file__,
                       'sender': param_dict['mail_sender'],
                       'recipients': param_dict['admin_email_address'],
                       'message_body_type': 'plain'
                       })
        return

    layer_info = pd.DataFrame(service_info['layers'] + service_info['tables']).set_index('id')
    layers = layer_info.index.tolist()

    # last_download_time is stored in local (AK) time but datetimes are stored in UTC in feature layer
    last_download_time_utc = TIMEZONE.localize(pd.to_datetime(last_download_time))
    last_download_time_utc -= last_download_time_utc.utcoffset()

    try:
        result = download.query_after_timestamp(service_url, token, layers, last_download_time_utc, ssl_cert)
    except:
        EMAILS.append({'message_body': 'An unexpected error occurred while querying submission info:\n%s'
                                       % traceback.format_exc(),
                       'subject': 'Error while running %s' % __file__,
                       'sender': param_dict['mail_sender'],
                       'recipients': param_dict['admin_email_address'],
                       'message_body_type': 'plain'
                       })
        return

    submissions = pd.DataFrame([feature['attributes'] for feature in result['layers'][0]['features']]) # 1st layer = survey123 parent table
    if not len(submissions):
        log_and_exit(log_info, [log_dir, download_dir, timestamp, False, 'No new data to download'])

    field_types = {}
    for layer in result['layers']:
        if 'fields' in layer:
            field_types[layer_info.loc[layer['id'], 'name']] = \
                pd.Series({field['name']: field['type'] for field in layer['fields']})
    last_submission_time = datetime.fromtimestamp(
        submissions.EditDate.max() / 1000.0)  # timestamps are in local tz, surprisingly

    # If the last_sumbission_time was before the last_download_time, these data have already been downloaded
    if last_submission_time < datetime.strptime(last_download_time, TIMESTAMP_FORMAT):
        log_and_exit(log_info, [log_dir, download_dir, timestamp, False, 'No submission since last download'])

    # If the last submission time was within SUBMISSION_TICKET_INTERVAL, exit in case a user is making multiple
    #   submission at once that should be grouped into a single ticket. It's possible that two different users could
    #   be submitting within the SUBMISSION_TICKET_INTERVAL of one another, but it's not worth the overhead and
    #   management to download data submitted by one user and not another. Their submissions will be assigned different
    #   ticket numbers, but just wait until there's a gap of all submissions of SUBMISSION_TICKET_INTERVAL.
    if (datetime.now() - last_submission_time).seconds < SUBMISSION_TICKET_INTERVAL:
        log_and_exit(log_info, [log_dir, download_dir, timestamp, False,
                                'Last submission was within %d minutes' % (SUBMISSION_TICKET_INTERVAL / 60)])  # '''

    # download the attachments
    feature_service_name = os.path.basename(os.path.dirname(service_url)) if not service_info['description'] else \
    service_info['description']
    sqlite_path = os.path.join(download_dir,
                               '{}_{}.db'.format(feature_service_name, datetime.now().strftime('%Y%m%d-%H%M%S')))
    try:
        attachments = download.download_attachments(submissions.globalid, token, 0, layer_info, service_url,
                                                    download_dir, ssl_cert=ssl_cert,
                                                    additional_info={'sqlite_path': sqlite_path}, overwrite=False)
    except:
        EMAILS.append({'message_body': 'An unexpected error occurred while downloading attachments:\n%s'
                                       % traceback.format_exc(),
                       'subject': 'Error while running %s' % __file__,
                       'sender': param_dict['mail_sender'],
                       'recipients': param_dict['admin_email_address'],
                       'message_body_type': 'plain'
                       })
        return

    # Write data to SQLite db so everything can be stored as a single file
    engine = create_engine('sqlite:///' + sqlite_path)
    tables = {}

    with engine.connect() as conn:
        for layer in result['layers']:
            layer_name = layer_info.loc[layer['id'], 'name']
            # If features is empty
            data = pd.DataFrame(
                    [feature['attributes'] for feature in layer['features']]
            )
            if len(data):
                data.to_sql(layer_name, conn, index=False)
            tables[layer_name] = data
        if len(attachments):
            attachments.to_sql('attachments', conn, index=False)

    # Get operator codes from usernames/submitter
    operator_codes = db_utils.get_lookup_table(table='operators', index_col='code', value_col='agol_username',
                                               conn=landings_conn)

    # Create separate tickets for each user for both landing and track data
    submissions['submission_time'] = [datetime.fromtimestamp(round(ts / 1000)) for _, ts in
                                      submissions['CreationDate'].iteritems()]

    # There was a brief time when Survey123 stopped recognizing a logged in NPS user. I added a field to
    #   enter an NPS email to resolve this, so use that wherever the Creator field (which is automatically
    #   filled in by AGOL when a user is logged in) is blank
    if 'tracks_nps_email' in submissions:
        creator_blank_mask = submissions.Creator.isna() | (submissions.Creator == '')
        submissions.loc[creator_blank_mask, 'Creator'] = submissions.loc[creator_blank_mask, 'tracks_nps_email']

    submissions.rename(columns={'Creator': 'submitter'}, inplace=True)
    submitter_null_mask = (submissions.submitter == '') | submissions.submitter.isna()
    submissions.loc[submitter_null_mask, 'submitter'] = [operator_codes[o] for _, o in submissions.loc[
        submitter_null_mask, 'operator'].iteritems()]

    connection_dict = {'tracks': tracks_conn, 'landings': landings_conn}
    submissions = pd.concat([get_ticket(g, connection_dict)
                             for _, g in submissions.groupby(['submitter', 'submission_type'])])

    # Process and import the landing data. The flight tracks will need to be unzipped, but they need to be
    #   validated/edited before they can be imported. This is handled manually via a separate web app
    # Add the ticket column to the sqlite flights table
    for table_name, data_types in field_types.items():
        df = tables[table_name]
        for field, _ in data_types.loc[data_types == 'esriFieldTypeDate'].iteritems():
            df.loc[~df[field].isnull(), field] = [datetime.fromtimestamp(ts / 1000.0) if ts else pd.NaT for ts in
                                                  df.loc[~df[field].isnull(), field]]
        tables[table_name] = df

    all_flights = tables['flights']
    merged = all_flights.merge(submissions, on='globalid')
    all_flights['ticket'] = merged.ticket
    all_flights['submission_time'] = merged.submission_time
    all_landings = tables['landing_repeat']

    # Check for any flights with landings submitted via Excel
    all_flights['is_from_excel'] = False
    excel_landing_flights = all_flights.loc[all_flights.landings_submission_type == 'excel']\
        .merge(attachments, left_on='globalid', right_on='parentglobalid')
    if len(excel_landing_flights):
        for _, info in excel_landing_flights.iterrows():
            excel_path = os.path.join(download_dir, 'attachments', info['name'])
            excel_flights = []

            try:
                excel_flights, excel_landings = process_excel_submission(
                    excel_path,
                    landings_conn,
                    info,
                    param_dict,
                    all_flights.columns
                )
            except Exception as e:
                html_li = '<li>The Excel file {filename} could not be processed because {error}</li>'\
                    .format(filename=info['name'], error=e)
                MESSAGES.append(
                    {'ticket': info.ticket, 'message': html_li, 'recipients': param_dict['admin_email_address'],
                     'type': 'landings', 'level': 'error'}
                )
                ERRORS.append(traceback.format_exc())

                # Try to remove the invalid submission
                try:
                    failed_object_ids = delete_from_feature_service([info.parentglobalid], token, service_url, ssl_cert)
                    if len(failed_object_ids):
                        html_li = ('<li>Unable to delete feature with Object ID {object_ids}</li>') \
                            .format(object_ids=', '.join(failed_object_ids.astype(str)))
                        MESSAGES.append(
                            {'ticket': info.ticket, 'message': html_li, 'recipients': param_dict['admin_email_address'],
                             'type': 'landings', 'level': 'error'})
                        ERRORS.append(traceback.format_exc())
                except Exception as e:
                    html_li = (
                        '<li>Could not delete AGOL ID {agol_id} because "{error}". This error occurred after failing to process {filename}</li>') \
                        .format(agol_id=info.parentglobalid, error=e, filename=info['name'])
                    MESSAGES.append({'ticket': info.ticket, 'message': html_li, 'recipients': param_dict['admin_email_address'],
                                     'type': 'landings', 'level': 'error'})
                    ERRORS.append(traceback.format_exc())

            # If there were errors, the flights and landing DataFrames will be empty
            if len(excel_flights):
                excel_landings.CreationDate = info.CreationDate
                all_flights = pd.concat([all_flights, excel_flights], ignore_index=True)
                all_landings = pd.concat([all_landings, excel_landings], ignore_index=True)

            # Remove the flight record that just had the Excel file because it
            all_flights = all_flights.loc[all_flights.objectid != info.objectid]

    # For some reason Pandas v 1.x doesn't read the departure_datetime column as Pandas datetimes even though they're
    #   all datetime.Datetimes. Instead, they're objects so explicitly cast them as datetimes
    all_flights.landing_datetime = pd.to_datetime(all_flights.landing_datetime)

    # Get operator emails from landings DB
    operator_emails = db_utils.get_lookup_table(table='operators', index_col='agol_username', value_col='email',
                                                conn=landings_conn)
    param_dict['agol_users'] = operator_emails
    submitter_emails = submissions.groupby(['ticket', 'submission_type']).first() \
        .apply(lambda r: get_submitter_email(r.submitter, param_dict['agol_users']), axis=1)

    submissions['tracks_operator'] = submissions.loc[submissions.submission_type == 'tracks', 'operator'].fillna('NPS')
    all_flights['tracks_operator'] = all_flights.loc[all_flights.submission_type == 'tracks', 'operator']
    all_flights['landing_operator'] = all_flights.loc[all_flights.submission_type == 'landings', 'operator']
    # The submitter being logged in *and* the username ending in _nps is the only way any non-operator could submit
    #   data. The operator field will be blank though, so fill it in
    all_flights.operator.fillna('NPS', inplace=True)

    # Add the ticket number to any attachment info files
    '''attachment_dir = os.path.join(download_dir, 'attachments')
    for attachment_info in all_flights.merge(attachments, left_on='globalid', right_on='parentglobalid').iterrows():
        json_path = os.path.join(attachment_dir, attachment_info['name'])
        with open(json_path) as f:
            info = json.load(f)
        info['ticket'] = attachment_info.ticket
        with open(json_path, 'w') as f:
            json.dump(info, f)'''

    # If there were any landings submitted, import them into the database
    receipt_dir = os.path.join(log_dir, 'receipts')
    receipt_params = param_dict['landing_receipt']
    if (submissions.submission_type == 'landings').any():
        # Try to import 1 ticket at a time so submissions from all users aren't affected by an error from 1 submitter
        _html_receipt_columns = ['departure_datetime', 'scenic_route', 'registration', 'aircraft_type',
                                 'submission_time', 'objectid']
        for ticket, flights in all_flights.loc[all_flights.submission_type == 'landings'].groupby('ticket'):
            # Create a blank df if there were no landings at all (but there was at least still one flight)
            landings = all_landings.loc[all_landings.parentglobalid.isin(flights.globalid)] if len(all_landings) else pd.DataFrame()
            html_table = get_html_table(
                flights.rename(columns=LANDINGS_FLIGHT_COLUMNS).reindex(columns=_html_receipt_columns))
            if len(landings) == 0:
                html_li = ('<li>There were no landings submitted for the following flights:<br>{table}<br></li>') \
                    .format(table=html_table)
                MESSAGES.append(
                    {'ticket': ticket, 'message': html_li, 'recipients': param_dict['landing_data_stewards'],
                     'type': 'landings', 'level': 'warning'})

                # Try to delete these flights so they don't keep getting processed
                agol_ids = flights.globalid
                try:
                    delete_from_feature_service(agol_ids, token, service_url, ssl_cert)
                except Exception as e:
                    html_li = (
                        '<li>The following error occurred while trying to delete features with AGOL IDs {agol_ids}, "{error}", while processing the following flights:<br>{table}<br></li>') \
                        .format(agol_ids=', '.join(agol_ids.astype(str)), error=e, table=html_table)
                    MESSAGES.append(
                        {'ticket': ticket, 'message': html_li, 'recipients': param_dict['admin_email_address'],
                         'type': 'landings', 'level': 'error'})
                    ERRORS.append(traceback.format_exc())

                # Then go to the next ticket
                continue

            # If there are landings, try to import them
            try:
                excel_path, html_df, imported_flights = \
                    import_landings(flights, ticket, landings_conn, sqlite_path, landings, receipt_dir,
                                    receipt_params['template'], receipt_params['header_img'],
                                    receipt_params['sheet_password'], param_dict['landing_data_stewards'])
            except Exception as e:
                tb_frame = get_traceback_frame()
                html_li = (
                    '<li>An unexpected error, "{error}", occurred on line {lineno} of {script} while processing the following flights:<br>{table}<br></li>') \
                    .format(error=e, lineno=tb_frame.lineno, script=os.path.basename(tb_frame.filename),
                            table=html_table)
                MESSAGES.append(
                    {'ticket': ticket, 'message': html_li, 'recipients': param_dict['landing_data_stewards'],
                     'type': 'landings', 'level': 'error'})
                ERRORS.append(traceback.format_exc())
                continue  # '''

            if not excel_path:  # There weren't any new landings
                continue

            # Delete any flights that weren't submitted via Excel
            agol_ids = imported_flights.agol_global_id[~imported_flights.is_from_excel]
            if len(agol_ids):
                try:
                    failed_object_ids = delete_from_feature_service(agol_ids, token, service_url, ssl_cert)
                    if len(failed_object_ids):
                        html_li = ('<li>Unable to delete features with Object IDs {object_ids}<br></li>') \
                            .format(object_ids=', '.join(failed_object_ids.astype(str)), table=html_table)
                        MESSAGES.append(
                            {'ticket': ticket, 'message': html_li, 'recipients': param_dict['admin_email_address'],
                             'type': 'landings', 'level': 'error'})
                        ERRORS.append(traceback.format_exc())
                except Exception as e:
                    html_li = (
                        '<li>The following error occurred while trying to delete features with AGOL IDs {agol_ids}, "{error}", while processing the following flights:<br>{table}<br></li>') \
                        .format(agol_ids=', '.join(agol_ids.astype(str)), error=e, table=html_table)
                    MESSAGES.append({'ticket': ticket, 'message': html_li, 'recipients': param_dict['admin_email_address'],
                                     'type': 'landings', 'level': 'error'})
                    ERRORS.append(traceback.format_exc())

            start_datetime = all_landings['CreationDate'].min()
            end_datetime = datetime.now()
            if submitter_emails[ticket].squeeze():
                EMAILS.append({'message_body': get_email_html(html_df, start_datetime, end_datetime,
                                                              column_widths={'Landing Locations': '100px'}),
                               'subject': 'Scenic/air taxi landing submission receipt - ticket #%s' % ticket,
                               'sender': param_dict['mail_sender'],
                               'recipients': submitter_emails[ticket].squeeze(),
                               'attachments': [excel_path],
                               'message_body_type': 'html'
                               })

    # Pre-process track data
    track_submissions = submissions.loc[submissions.submission_type == 'tracks']
    if len(track_submissions):
        invalid_file_features = []
        for _, file_info in submissions.merge(attachments, left_on='globalid', right_on='parentglobalid').iterrows():
            name, ext = os.path.splitext(file_info['name'])

            if ext not in import_track.READ_FUNCTIONS and ext.lower() != '.zip':
                html_li = ('<li>The file "{file}" is in an invalid format. Accepted formats are {file_formats}.</li>').format(
                    file=file_info['name'], file_formats=', '.join(import_track.READ_FUNCTIONS))
                MESSAGES.append({'ticket': file_info['ticket'], 'message': html_li, 'recipients': param_dict['track_data_stewards'],
                                 'type': 'tracks', 'level': 'error'})
            invalid_file_features.append(file_info.parentglobalid)

        geojson_paths, submitted_files = prepare_track_data(param_dict, download_dir, tracks_conn, submissions)
        TRACK_JSON_FILES.extend(geojson_paths)

        try:
            failed_object_ids = delete_from_feature_service(submitted_files.index.tolist() + invalid_file_features, token, service_url, ssl_cert)
            if len(failed_object_ids):
                EMAILS.append({
                                  'message_body': 'Unable to delete the following features: %s. These should be manually deleted.' % failed_object_ids.astype(
                                      str),
                                  'subject': 'Error while attempting to delete features',
                                  'sender': param_dict['mail_sender'],
                                  'recipients': param_dict['admin_email_address'],
                                  'message_body_type': 'plain'
                                  })
        except Exception as e:
            EMAILS.append({
                              'message_body': 'Unable to delete the following features: %s. These should be manually deleted.' % submitted_files.index.tolist(),
                              'subject': 'Error while attempting to delete features',
                              'sender': param_dict['mail_sender'],
                              'recipients': param_dict['admin_email_address'],
                              'message_body_type': 'plain'
                              })

        # Compose email to track editors
        if len(geojson_paths):
            msg = ('There are new track data to edit and import into the overflights database as of {timestamp}. Go to '
                   '{url} in a Google Chrome window to view and edit the tracks.') \
                .format(url=param_dict['track_editor_url'], timestamp=timestamp)
            subject = 'New track data submitted on %s' % datetime.now().strftime('%b %d, %Y')
            recipients = param_dict['track_data_stewards']
            EMAILS.append({'message_body': msg,
                           'subject': subject,
                           'sender': param_dict['mail_sender'],
                           'recipients': recipients})

        track_submissions.set_index('globalid', inplace=True)
        track_submissions['Filename'] = submitted_files
        for ticket, s in track_submissions.groupby('ticket'):
            message_info = get_track_receipt_email(s, ticket, param_dict['mail_sender'], param_dict['agol_users'])
            if len(message_info):
                EMAILS.append(message_info)
                global DATA_PROCESSED
                DATA_PROCESSED = True
        # EMAILS.extend([get_track_receipt_email(s, ticket, param_dict['mail_sender'], param_dict['agol_users']) for ticket, s in track_submissions.groupby('ticket')])

    # Reduce submissions df to one per ticket
    if len(MESSAGES):
        tickets = submissions.sort_values('submission_time') \
            .groupby(['ticket', 'submission_type']) \
            .first() \
            .reset_index()
        EMAILS.extend(compose_error_notifications(tickets, param_dict))


def send_email_per_recipient(message_params):
    '''
    Helper function to send mail per recipient given in message_params['recipients'] because our homegrown mail server
    sometimes seems to skip messages to some recipients when a list of recipeints is given
    :param args: List of positional args for process_emails.send_mail()
    :param message_params: Dict of kwargs for process_emails.send_mail()
    :return: a list of messages that resulted in errors
    '''

    # If the recipient isn't a list of recipients, make it one so it can be iterated over
    recipients = message_params['recipients']
    if isinstance(recipients, six.string_types):
        recipients = [recipients]

    failed_mail = []
    for address in recipients:
        params = message_params.copy()
        params['recipients'] = [address]  # send_mail() still expects an iterable
        try:
            process_emails.send_email(**params)
        except Exception as e:
            params['error'] = traceback.format_exc()
            failed_mail.append(params)
        time.sleep(1)  # not sure if the issue is that messages are sent at the exact same time, but this seems to work

    return failed_mail


def main(config_json):
    '''
    Ensure that a log file is always written (unless the error occurs in write_log(), which is unlikely)
    '''
    global EMAILS

    timestamp = datetime.now().strftime(TIMESTAMP_FORMAT)

    try:
        params = read_json_params(config_json)
        # Open connections to the DBs and begin transactions so that if there's an exception, no data are inserted
        connection_info = params['db_credentials']
        connection_template = 'postgresql://{username}:{password}@{ip_address}:{port}/{db_name}'
        landings_engine = create_engine(connection_template.format(**connection_info['landings']))
        tracks_engine = create_engine(connection_template.format(**connection_info['tracks']))

        with landings_engine.connect() as l_conn, tracks_engine.connect() as t_conn, l_conn.begin(), t_conn.begin():
            poll_feature_service(params['log_dir'], params['download_dir'], params, params['ssl_cert'], l_conn, t_conn)

    except Exception as error:
        tb_frame = get_traceback_frame()
        exc_type, exc_value, exc_traceback = sys.exc_info()
        traceback_details = {'lineno': tb_frame.lineno,
                             'filename': tb_frame.filename,
                             'exception_type': exc_type.__name__,
                             'exception_message': str(error),
                             'full_traceback': traceback.format_exc()
                             }
        EMAILS.append({'message_body': 'An unexpected error occurred while polling the flight data feature layer or '
                                       'processing submissions:\n%s' % traceback_details['full_traceback'],
                       'subject': 'Error while running %s' % __file__,
                       'sender': params['mail_sender'],
                       'recipients': params['admin_email_address'],
                       'message_body_type': 'plain'
                       })

        # Remove any of the track JSONs that were created while processing these tickets. Since all database operations
        #   are run within a transaction, when there's an error, the transaction gets rolled back. That includes any
        #   tickets created in the submissions table, which need to be present when a track is imported. In order for
        #   the available track files to stay in sync with the database, all track files associated with those rolled
        #   back tickets should be deleted
        for f in TRACK_JSON_FILES:
            try:
                os.remove(f)
            except:
                continue
    # '''

    # send emails here so that all DB transactions are committed before actually sending anything
    # still wrap in try/except block so if any emails fail, someone can be notified (and the log can be amended)
    # Start the mail server
    server = smtplib.SMTP(params['mail_server_credentials']['server_name'], params['mail_server_credentials']['port'])
    server.starttls()
    server.ehlo()  # '''

    script_name = os.path.basename(__file__).rstrip('.py')
    log_info = read_logs(params['log_dir'], script_name)
    previous_emails = pd.DataFrame(
        log_info['emails'].explode().dropna().tolist()) if 'emails' in log_info else pd.DataFrame()
    if len(previous_emails):
        try:
            previous_emails['recipients'] = previous_emails.recipients.apply(sorted).astype(str)
        except:
            pass

    failed_emails = []

    for msg_info in EMAILS:  # + steward_notifications:
        # If this message has not been sent before, send it. Don't use the subject line to determine this because it
        #   contains a ticket number, which changes every time there's an attempt to process the data
        sent = False
        if len(previous_emails):
            try:
                sent = ((previous_emails.message_body == msg_info['message_body']) &
                        (previous_emails.recipients.str.contains(str(sorted(msg_info['recipients'])), regex=False))
                        ).any()
            except:
                pass

        if not sent:
            msg_info['server'] = server
            failed = send_email_per_recipient(msg_info)
            failed_emails.extend(failed)

    # Write the log file
    log_file, log = write_log(params['log_dir'], params['download_dir'], timestamp, DATA_PROCESSED)

    # If any failed, add them to the log file. The log file isn't actually created until write_log is called, so do this after
    if len(failed_emails):
        log['failed_emails'] = [{k: v for k, v in e.items() if k != 'server'} for e in failed_emails]
        with open(log_file, 'w') as j:
            json.dump(log, j, indent=4)


if __name__ == '__main__':
    sys.exit(main(*sys.argv[1:]))
